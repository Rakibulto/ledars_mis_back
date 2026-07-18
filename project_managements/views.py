from calendar import monthrange
from io import BytesIO

from django.db.models import Prefetch, Q
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.text import slugify
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from rest_framework import permissions, status, viewsets
from rest_framework.decorators import action
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend

from inventory.views.core import CreatedByMixin
from paginations import Pagination
from .services.dashboard_service import build_project_management_dashboard_payload
from .services.expense_export_service import build_project_management_expense_pdf_bytes

from .models import (
    Advance,
    ProjectManagementExpense,
    ProjectManagementPlanAttachment,
    ProjectManagementPlanSubPlan,
    ProjectManagementPlanWorkItem,
    ProjectManagementProject,
    ProjectManagementProjectPlan,
    ProjectManagementUnit,
)
from .serializers import (
    AdvanceSerializer,
    ProjectManagementExpenseSerializer,
    ProjectManagementPlanAttachmentSerializer,
    ProjectManagementPlanWorkItemSerializer,
    ProjectManagementPlanSerializer,
    ProjectManagementProjectSerializer,
    ProjectManagementUnitSerializer,
    ProjectOverviewSerializer,
)


def _resolve_overview_date_range(period, start_date_str, end_date_str):
    today = timezone.localdate()
    range_start = None
    range_end = None

    if period == "daily":
        range_start = range_end = today
    elif period == "monthly":
        range_start = today.replace(day=1)
        range_end = today.replace(day=monthrange(today.year, today.month)[1])
    elif period == "yearly":
        range_start = today.replace(month=1, day=1)
        range_end = today.replace(month=12, day=31)

    if start_date_str:
        parsed = parse_date(start_date_str)
        if parsed:
            range_start = parsed
    if end_date_str:
        parsed = parse_date(end_date_str)
        if parsed:
            range_end = parsed

    return range_start, range_end


def _overview_activity_date_q(range_start, range_end):
    """Match activities whose deliverable (end_date) or start_date falls in range."""
    if not range_start and not range_end:
        return Q()

    if range_start and range_end:
        return Q(end_date__gte=range_start, end_date__lte=range_end) | Q(
            end_date__isnull=True,
            start_date__gte=range_start,
            start_date__lte=range_end,
        )
    if range_start:
        return Q(end_date__gte=range_start) | Q(
            end_date__isnull=True, start_date__gte=range_start
        )
    return Q(end_date__lte=range_end) | Q(
        end_date__isnull=True, start_date__lte=range_end
    )


class ProjectManagementUnitViewSet(CreatedByMixin, viewsets.ModelViewSet):
    serializer_class = ProjectManagementUnitSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = Pagination
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ["status"]
    search_fields = ["name", "description"]
    ordering_fields = ["name", "created_at", "updated_at", "status"]
    ordering = ["name"]

    def get_queryset(self):
        return ProjectManagementUnit.objects.select_related("created_by").all()


class ProjectManagementDashboardViewSet(viewsets.ViewSet):
    permission_classes = [permissions.IsAuthenticated]

    def list(self, request):
        return Response(build_project_management_dashboard_payload())


class ProjectManagementProjectViewSet(CreatedByMixin, viewsets.ModelViewSet):
    serializer_class = ProjectManagementProjectSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = Pagination
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    # project_type / implementation_type are JSONField lists — not auto-filterable
    filterset_fields = [
        "status",
        "reporting_frequency",
        "risk_level",
        "donor",
        "project_manager",
    ]
    search_fields = [
        "code",
        "title",
        "short_name",
        "location",
        "sector",
        "background",
        "objectives",
        "donor__name",
        "project_manager__username",
    ]
    ordering_fields = [
        "created_at",
        "updated_at",
        "title",
        "start_date",
        "end_date",
        "budget_amount",
        "status",
    ]
    ordering = ["-created_at"]

    def get_queryset(self):
        return ProjectManagementProject.objects.select_related(
            "donor", "project_manager", "created_by", "materials_expense"
        ).prefetch_related(
            "assigned_users",
            "plans__approved_by",
            "plans__assigned_users",
            "plans__sub_plans__assigned_users",
            "plans__work_items__assigned_to",
            "plans__work_items__approved_by",
            "plans__attachments__uploaded_by",
            "materials__plan",
        )

    @action(detail=True, methods=["get"], url_path="export-roadmap-excel")
    def export_roadmap_excel(self, request, pk=None):
        project = self.get_object()
        file_bytes = build_project_management_roadmap_workbook(project)
        file_stub = slugify(project.code or project.title or f"project-{project.id}") or f"project-{project.id}"

        response = HttpResponse(
            file_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{file_stub}-roadmap.xlsx"'
        return response

    @action(detail=False, methods=["get"], url_path="options")
    def options(self, request):
        """Lightweight id/title list for filter dropdowns."""
        qs = ProjectManagementProject.objects.only("id", "title", "code").order_by("title")
        return Response(
            [{"id": item.id, "title": item.title, "code": item.code} for item in qs]
        )

    @action(detail=False, methods=["get"], url_path="overview")
    def overview(self, request):
        """
        Lightweight project activity overview for large datasets.
        Query params:
          - project: project id
          - start_date / end_date: YYYY-MM-DD
          - period: daily | monthly | yearly
          - pagination=true&page=&page_size=
        """
        project_id = request.query_params.get("project")
        period = (request.query_params.get("period") or "").strip().lower()
        range_start, range_end = _resolve_overview_date_range(
            period,
            request.query_params.get("start_date"),
            request.query_params.get("end_date"),
        )
        date_q = _overview_activity_date_q(range_start, range_end)

        matching_plan_ids = None
        matching_sub_ids = None
        if date_q:
            matching_sub_ids = list(
                ProjectManagementPlanSubPlan.objects.filter(date_q).values_list(
                    "id", flat=True
                )
            )
            matching_plan_ids = list(
                ProjectManagementProjectPlan.objects.filter(date_q).values_list(
                    "id", flat=True
                )
            )
            parent_ids_from_subs = list(
                ProjectManagementPlanSubPlan.objects.filter(
                    id__in=matching_sub_ids
                ).values_list("plan_id", flat=True)
            )
            plan_id_set = set(matching_plan_ids) | set(parent_ids_from_subs)

            sub_plan_qs = (
                ProjectManagementPlanSubPlan.objects.filter(
                    Q(id__in=matching_sub_ids) | Q(plan_id__in=matching_plan_ids)
                )
                .prefetch_related("assigned_users")
                .order_by("sort_order", "id")
            )
            plan_qs = (
                ProjectManagementProjectPlan.objects.filter(id__in=plan_id_set)
                .prefetch_related(
                    "assigned_users",
                    "work_items__assigned_to",
                    "work_items__approved_by",
                    Prefetch("sub_plans", queryset=sub_plan_qs),
                )
                .order_by("serial_no", "id")
            )
        else:
            sub_plan_qs = ProjectManagementPlanSubPlan.objects.prefetch_related(
                "assigned_users"
            ).order_by("sort_order", "id")
            plan_qs = ProjectManagementProjectPlan.objects.prefetch_related(
                "assigned_users",
                "work_items__assigned_to",
                "work_items__approved_by",
                Prefetch("sub_plans", queryset=sub_plan_qs),
            ).order_by("serial_no", "id")

        qs = ProjectManagementProject.objects.only(
            "id", "code", "title", "short_name", "status", "start_date", "end_date"
        ).prefetch_related("assigned_users", Prefetch("plans", queryset=plan_qs))

        if project_id:
            qs = qs.filter(id=project_id)

        if date_q:
            qs = qs.filter(
                Q(plans__id__in=matching_plan_ids)
                | Q(plans__sub_plans__id__in=matching_sub_ids)
            ).distinct()

        qs = qs.order_by("title", "id")

        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = ProjectOverviewSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = ProjectOverviewSerializer(qs, many=True)
        return Response(serializer.data)


class ProjectManagementPlanViewSet(viewsets.ModelViewSet):
    serializer_class = ProjectManagementPlanSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = Pagination
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ["project", "status"]
    search_fields = ["title", "description", "project__title"]
    ordering_fields = ["serial_no", "start_date", "end_date", "created_at"]
    ordering = ["project", "serial_no"]

    def get_queryset(self):
        return ProjectManagementProjectPlan.objects.select_related("project", "approved_by").prefetch_related(
            "assigned_users",
            "sub_plans__assigned_users",
            "work_items__assigned_to",
            "work_items__approved_by",
            "attachments__uploaded_by",
        )

    @action(detail=True, methods=["post"])
    def approve(self, request, pk=None):
        plan = self.get_object()

        if plan.status != "Completed":
            return Response(
                {"detail": "Only completed task plans can be approved."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if plan.approval_status != "Approved":
            plan.mark_approved(request.user)

        serializer = self.get_serializer(plan)
        return Response(serializer.data)


class ProjectManagementPlanAttachmentViewSet(viewsets.ModelViewSet):
    serializer_class = ProjectManagementPlanAttachmentSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = Pagination
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ["plan", "work_item"]
    ordering_fields = ["created_at", "id"]
    ordering = ["-created_at"]

    def get_queryset(self):
        return ProjectManagementPlanAttachment.objects.select_related(
            "plan",
            "work_item",
            "uploaded_by",
            "plan__project",
        )


class ProjectManagementPlanWorkItemViewSet(viewsets.ModelViewSet):
    serializer_class = ProjectManagementPlanWorkItemSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = Pagination
    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_fields = ["plan", "state", "approval_status"]
    ordering_fields = ["sort_order", "scheduled_date", "created_at", "updated_at"]
    ordering = ["sort_order", "id"]

    def get_queryset(self):
        return ProjectManagementPlanWorkItem.objects.select_related(
            "plan", "plan__project", "assigned_to", "approved_by"
        ).prefetch_related("attachments")

    @action(detail=True, methods=["post"])
    def approve(self, request, pk=None):
        work_item = self.get_object()

        if work_item.state != "Done":
            return Response(
                {"detail": "Only completed execution tasks can be approved."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if work_item.approval_status != "Approved":
            work_item.mark_approved(request.user)

        serializer = self.get_serializer(work_item)
        return Response(serializer.data)


def build_project_management_expense_workbook(expenses, include_items=True):
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Expenses"

    title_fill = PatternFill("solid", fgColor="1F4E78")
    title_font = Font(color="FFFFFF", bold=True)

    headers = [
        "Invoice",
        "Project",
        "Task",
        "Title",
        "Vendor",
        "Expense Date",
        "Status",
        "Currency",
        "Total Amount",
        "Description",
    ]
    summary_sheet.append(headers)

    for cell in summary_sheet[1]:
        cell.fill = title_fill
        cell.font = title_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for expense in expenses:
        summary_sheet.append(
            [
                expense.invoice_number,
                expense.project.title,
                expense.plan.title if expense.plan else "",
                expense.title,
                expense.vendor_name,
                expense.expense_date.isoformat() if expense.expense_date else "",
                expense.status,
                expense.currency,
                float(expense.total_amount or 0),
                expense.description,
            ]
        )

    summary_sheet.freeze_panes = "A2"
    summary_sheet.column_dimensions["A"].width = 16
    summary_sheet.column_dimensions["B"].width = 28
    summary_sheet.column_dimensions["C"].width = 24
    summary_sheet.column_dimensions["D"].width = 28
    summary_sheet.column_dimensions["E"].width = 22
    summary_sheet.column_dimensions["F"].width = 14
    summary_sheet.column_dimensions["G"].width = 14
    summary_sheet.column_dimensions["H"].width = 12
    summary_sheet.column_dimensions["I"].width = 16
    summary_sheet.column_dimensions["J"].width = 44

    if include_items:
        items_sheet = workbook.create_sheet("Expense Items")
        item_headers = [
            "Invoice",
            "Item",
            "Description",
            "Quantity",
            "Unit Price",
            "Line Total",
        ]
        items_sheet.append(item_headers)

        for cell in items_sheet[1]:
            cell.fill = title_fill
            cell.font = title_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for expense in expenses:
            for item in expense.items.all():
                items_sheet.append(
                    [
                        expense.invoice_number,
                        item.title,
                        item.description,
                        float(item.quantity or 0),
                        float(item.unit_price or 0),
                        float(item.line_total or 0),
                    ]
                )

        items_sheet.freeze_panes = "A2"
        items_sheet.column_dimensions["A"].width = 16
        items_sheet.column_dimensions["B"].width = 28
        items_sheet.column_dimensions["C"].width = 36
        items_sheet.column_dimensions["D"].width = 12
        items_sheet.column_dimensions["E"].width = 16
        items_sheet.column_dimensions["F"].width = 16

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def build_project_management_roadmap_workbook(project):
    workbook = Workbook()

    title_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    accent_fill = PatternFill("solid", fgColor="EAF4EA")
    title_font = Font(color="FFFFFF", bold=True)
    section_font = Font(bold=True, color="0F172A")
    wrap_alignment = Alignment(vertical="top", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    plans = list(project.plans.all())
    all_work_items = []
    attachment_rows = []

    for plan in plans:
        work_items = list(plan.work_items.order_by("sort_order", "id"))
        all_work_items.extend([(plan, item) for item in work_items])

        for attachment in plan.attachments.all():
            attachment_rows.append(
                [
                    plan.serial_no,
                    plan.title,
                    attachment.work_item.title if attachment.work_item else "Plan",
                    attachment.display_name,
                    attachment.original_name or "",
                    attachment.uploaded_by.username if attachment.uploaded_by else "",
                    attachment.created_at.strftime("%Y-%m-%d %H:%M") if attachment.created_at else "",
                ]
            )

    completed_plans = sum(1 for plan in plans if plan.status == "Completed")
    active_plans = sum(1 for plan in plans if plan.status == "In Progress")
    on_hold_plans = sum(1 for plan in plans if plan.status == "On Hold")
    pending_plans = len(plans) - completed_plans - active_plans - on_hold_plans
    completed_work_items = sum(1 for _, item in all_work_items if item.state == "Done")
    active_work_items = sum(1 for _, item in all_work_items if item.state == "Doing")
    pending_work_items = len(all_work_items) - completed_work_items - active_work_items

    overview_sheet = workbook.active
    overview_sheet.title = "Overview"
    overview_sheet.merge_cells("A1:F1")
    overview_sheet["A1"] = f"Project Roadmap Export - {project.title}"
    overview_sheet["A1"].fill = title_fill
    overview_sheet["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    overview_sheet["A1"].alignment = center_alignment

    overview_rows = [
        ("Project Code", project.code or "—", "Status", project.status or "—"),
        (
            "Project Title",
            project.title or "—",
            "Project Type",
            ", ".join(project.project_type or []) or "—",
        ),
        (
            "Donor",
            project.donor.name if project.donor else "—",
            "Implementation",
            ", ".join(project.implementation_type or []) or "—",
        ),
        ("Project Manager", project.project_manager.username if project.project_manager else "—", "Risk Level", project.risk_level or "—"),
        ("Start Date", project.start_date.isoformat() if project.start_date else "—", "End Date", project.end_date.isoformat() if project.end_date else "—"),
        ("Assigned Team", ", ".join(user.username for user in project.assigned_users.all()) or "—", "Reporting", project.reporting_frequency or "—"),
    ]

    overview_sheet.append([])
    overview_sheet.append(["Field", "Value", "Field", "Value"])
    for cell in overview_sheet[3]:
        cell.fill = section_fill
        cell.font = section_font
        cell.alignment = center_alignment

    for row in overview_rows:
        overview_sheet.append(list(row))

    summary_start = overview_sheet.max_row + 2
    overview_sheet.append(["Roadmap Summary", "", "", ""])
    for cell in overview_sheet[summary_start]:
        cell.fill = accent_fill
        cell.font = section_font
        cell.alignment = wrap_alignment

    overview_sheet.append(["Plans", len(plans), "Work Items", len(all_work_items)])
    overview_sheet.append(["Completed Plans", completed_plans, "Completed Work Items", completed_work_items])
    overview_sheet.append(["Active Plans", active_plans, "Active Work Items", active_work_items])
    overview_sheet.append(["On Hold Plans", on_hold_plans, "Pending Work Items", pending_work_items])
    overview_sheet.append(["Pending Plans", pending_plans, "Attachments", len(attachment_rows)])
    overview_sheet.append([])
    overview_sheet.append(["Objectives", project.objectives or "—"])
    overview_sheet.append(["Expected Outcomes", project.expected_outcomes or "—"])
    overview_sheet.append(["Background", project.background or "—"])
    overview_sheet.append(["Monitoring Plan", project.monitoring_plan or "—"])
    overview_sheet.append(["Notes", project.notes or "—"])

    for column in ["A", "B", "C", "D", "E", "F"]:
        overview_sheet.column_dimensions[column].width = 24

    roadmap_sheet = workbook.create_sheet("Roadmap")
    roadmap_headers = [
        "SL",
        "Task / Plan",
        "Description",
        "Status",
        "Approval",
        "Start Date",
        "End Date",
        "Duration (Days)",
        "Assigned Team",
        "Work Items",
        "Completed",
        "In Progress",
        "Pending",
        "Plan Attachments",
    ]
    roadmap_sheet.append(roadmap_headers)
    for cell in roadmap_sheet[1]:
        cell.fill = title_fill
        cell.font = title_font
        cell.alignment = center_alignment

    for plan in plans:
        work_items = list(plan.work_items.order_by("sort_order", "id"))
        plan_attachments = [attachment.display_name for attachment in plan.attachments.all() if not attachment.work_item_id]
        roadmap_sheet.append(
            [
                plan.serial_no,
                plan.title,
                plan.description or "",
                plan.status,
                f"{plan.approval_status}{f' • {plan.approved_by.username}' if plan.approved_by else ''}",
                plan.start_date.isoformat() if plan.start_date else "",
                plan.end_date.isoformat() if plan.end_date else "",
                plan.duration_days,
                ", ".join(user.username for user in plan.assigned_users.all()) or "Unassigned",
                len(work_items),
                sum(1 for item in work_items if item.state == "Done"),
                sum(1 for item in work_items if item.state == "Doing"),
                sum(1 for item in work_items if item.state == "Todo"),
                "\n".join(plan_attachments) or "—",
            ]
        )

    roadmap_sheet.freeze_panes = "A2"
    roadmap_widths = {
        "A": 8,
        "B": 30,
        "C": 44,
        "D": 16,
        "E": 24,
        "F": 14,
        "G": 14,
        "H": 16,
        "I": 28,
        "J": 12,
        "K": 12,
        "L": 12,
        "M": 12,
        "N": 28,
    }
    for column, width in roadmap_widths.items():
        roadmap_sheet.column_dimensions[column].width = width

    details_sheet = workbook.create_sheet("Work Item Details")
    detail_headers = [
        "Plan SL",
        "Plan Title",
        "Plan Status",
        "Work Item Order",
        "Work Item",
        "State",
        "Assigned To",
        "Scheduled Start",
        "Scheduled End",
        "Completed At",
        "Approval Status",
        "Approved By",
        "Approved At",
        "Notes",
        "Issues / Blockers",
        "Attachments",
    ]
    details_sheet.append(detail_headers)
    for cell in details_sheet[1]:
        cell.fill = title_fill
        cell.font = title_font
        cell.alignment = center_alignment

    for plan in plans:
        work_items = list(plan.work_items.order_by("sort_order", "id"))
        if not work_items:
            details_sheet.append(
                [
                    plan.serial_no,
                    plan.title,
                    plan.status,
                    "",
                    "No execution items added",
                    "",
                    "",
                    "",
                    "",
                    "",
                    plan.approval_status,
                    plan.approved_by.username if plan.approved_by else "",
                    plan.approved_at.strftime("%Y-%m-%d %H:%M") if plan.approved_at else "",
                    plan.description or "",
                    "",
                    "",
                ]
            )
            continue

        for item in work_items:
            item_attachments = [attachment.display_name for attachment in item.attachments.all()]
            details_sheet.append(
                [
                    plan.serial_no,
                    plan.title,
                    plan.status,
                    item.sort_order,
                    item.title,
                    item.state,
                    item.assigned_to.username if item.assigned_to else "Unassigned",
                    item.scheduled_date.isoformat() if item.scheduled_date else "",
                    item.scheduled_end_date.isoformat() if item.scheduled_end_date else "",
                    item.completed_at.strftime("%Y-%m-%d %H:%M") if item.completed_at else "",
                    item.approval_status,
                    item.approved_by.username if item.approved_by else "",
                    item.approved_at.strftime("%Y-%m-%d %H:%M") if item.approved_at else "",
                    item.notes or "",
                    item.issues or "",
                    "\n".join(item_attachments) or "—",
                ]
            )

    details_sheet.freeze_panes = "A2"
    details_widths = {
        "A": 10,
        "B": 28,
        "C": 16,
        "D": 14,
        "E": 32,
        "F": 12,
        "G": 20,
        "H": 14,
        "I": 14,
        "J": 20,
        "K": 18,
        "L": 18,
        "M": 20,
        "N": 42,
        "O": 42,
        "P": 28,
    }
    for column, width in details_widths.items():
        details_sheet.column_dimensions[column].width = width

    attachments_sheet = workbook.create_sheet("Attachments")
    attachments_headers = [
        "Plan SL",
        "Plan Title",
        "Related To",
        "Attachment Name",
        "Original Name",
        "Uploaded By",
        "Uploaded At",
    ]
    attachments_sheet.append(attachments_headers)
    for cell in attachments_sheet[1]:
        cell.fill = title_fill
        cell.font = title_font
        cell.alignment = center_alignment

    for row in attachment_rows or [["", "", "No attachments", "", "", "", ""]]:
        attachments_sheet.append(row)

    attachments_sheet.freeze_panes = "A2"
    for column, width in {"A": 10, "B": 28, "C": 24, "D": 28, "E": 28, "F": 18, "G": 20}.items():
        attachments_sheet.column_dimensions[column].width = width

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.row == 1:
                    continue
                cell.alignment = wrap_alignment

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


class ProjectManagementExpenseViewSet(CreatedByMixin, viewsets.ModelViewSet):
    serializer_class = ProjectManagementExpenseSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = Pagination
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ["project", "plan", "status", "currency"]
    search_fields = [
        "invoice_number",
        "title",
        "description",
        "vendor_name",
        "project__title",
        "plan__title",
    ]
    ordering_fields = ["expense_date", "created_at", "updated_at", "total_amount", "invoice_number"]
    ordering = ["-expense_date", "-created_at"]

    def get_queryset(self):
        return ProjectManagementExpense.objects.select_related(
            "project", "plan", "created_by", "approved_by"
        ).prefetch_related("items")

    @action(detail=True, methods=["post"])
    def transition(self, request, pk=None):
        expense = self.get_object()
        next_status = request.data.get("status")

        if not next_status:
            return Response({"detail": "Status is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            expense.transition_status(next_status, request.user)
        except ValueError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        serializer = self.get_serializer(expense)
        return Response(serializer.data)

    @action(detail=True, methods=["get"], url_path="export-excel")
    def export_excel(self, request, pk=None):
        expense = self.get_object()
        file_bytes = build_project_management_expense_workbook([expense], include_items=True)

        response = HttpResponse(
            file_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{expense.invoice_number}.xlsx"'
        return response

    @action(detail=True, methods=["get"], url_path="export-pdf")
    def export_pdf(self, request, pk=None):
        expense = self.get_object()
        file_bytes = build_project_management_expense_pdf_bytes(expense)

        response = HttpResponse(file_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{expense.invoice_number}.pdf"'
        return response

    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_list_excel(self, request):
        expenses = self.filter_queryset(self.get_queryset())
        file_bytes = build_project_management_expense_workbook(expenses, include_items=True)

        response = HttpResponse(
            file_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="project-management-expenses.xlsx"'
        return response


class AdvanceViewSet(viewsets.ModelViewSet):
    serializer_class = AdvanceSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_queryset(self):
        return Advance.objects.select_related(
            "from_employee", "project"
        ).all()
